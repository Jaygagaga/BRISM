# 2015-02-02
import tweepy
import time
import pandas as pd
import os
from openpyxl import load_workbook
from TwitterData.twitter_authentication import bearer_token
# from twitter_authentication import bearer_token
client = tweepy.Client(bearer_token, wait_on_rate_limit=True)
user_dict = {}
place_dict = {}

# keyword1 = 'user_profile'
# keyword2 = 'silk road'
# author_profile = pd.DataFrame(columns = ['statuses_count', 'favourites_count'])
author_profile = pd.DataFrame(columns = ['username', 'name', 'profile_image_url', 'followers_count', 'following_count', 'listed_count',
                                 'tweet_count', 'description', 'location', 'url', 'verified', 'author_id'])
# silk_road =  pd.DataFrame(columns=['author_id','id','text','created_at','retweet_count','reply_count',
#                                      'like_count' ,'quote_count','geo','lang','conversation_id','referenced_tweets'])
#
#
bri=  pd.DataFrame(columns=['author_id','id','text','created_at','retweet_count','reply_count',
                                     'like_count' ,'quote_count','geo','lang','conversation_id','referenced_tweets'])
def scrapy(keyword1, keyword2=None,and_query =False, until_id = None,save_type = 'csv',start='',end='',
           file_path = '/Users/jie/phd_project/TwitterData/brism',user_filename='', tweet_filename = ''):
    i=0
    global author_profile
    global bri
    if until_id and and_query == True:
        pagination = tweepy.Paginator(client.search_all_tweets,
                                      query='"{}" {}'.format(keyword1, keyword2),
                                      user_fields=['username', 'id', 'name', 'public_metrics', 'description',
                                                   'location', 'profile_image_url',
                                                   'protected', 'url', 'verified'],
                                      tweet_fields=['id', 'created_at', 'geo', 'public_metrics', 'text', 'lang',
                                                    'referenced_tweets', 'conversation_id'],
                                      expansions=['author_id', 'referenced_tweets.id', 'geo.place_id'],
                                      # start_time = '2010-01-01T00:00:00Z',
                                      # end_time = '2016-01-01T00:00:00Z',
                                      until_id=until_id,
                                      max_results=500)
    if until_id and and_query ==False:
        pagination= tweepy.Paginator(client.search_all_tweets,
                                     query = '"{}"'.format(keyword1),
                                     user_fields = ['username','id', 'name', 'public_metrics', 'description', 'location', 'profile_image_url',
                                                    'protected', 'url', 'verified'],
                                     tweet_fields = ['id','created_at', 'geo', 'public_metrics', 'text','lang','referenced_tweets','conversation_id'],
                                     expansions = ['author_id','referenced_tweets.id', 'geo.place_id'],
                                     # start_time = '2010-01-01T00:00:00Z',
                                     # end_time = '2016-01-01T00:00:00Z',
                                     until_id =until_id,
                                     max_results=500)
    if not until_id and and_query ==True:
        pagination = tweepy.Paginator(client.search_all_tweets,
                                     query = '"{}" {}'.format(keyword1, keyword2),
                                     user_fields = ['username','id', 'name', 'public_metrics', 'description', 'location', 'profile_image_url',
                                                    'protected', 'url', 'verified'],
                                     tweet_fields = ['id','created_at', 'geo', 'public_metrics', 'text','lang','referenced_tweets','conversation_id'],
                                     expansions = ['author_id','referenced_tweets.id', 'geo.place_id'],
                                     start_time = '{}T00:00:00Z'.format(start),
                                     end_time = '{}T00:00:00Z'.format(end),
                                     # until_id =until_id,
                                  max_results=500)
    if not until_id and and_query ==False:
        pagination = tweepy.Paginator(client.search_all_tweets,
                                     query = '"{}"'.format(keyword1),
                                     user_fields = ['username','id', 'name', 'public_metrics', 'description', 'location', 'profile_image_url',
                                                    'protected', 'url', 'verified'],
                                     tweet_fields = ['id','created_at', 'geo', 'public_metrics', 'text','lang','referenced_tweets','conversation_id'],
                                     expansions = ['author_id','referenced_tweets.id', 'geo.place_id'],
                                     start_time = '{}T00:00:00Z'.format(start),
                                     end_time = '{}T00:00:00Z'.format(end),
                                     # until_id =until_id,
                                  max_results=500)
    for response in pagination:
        # print(response.data)
        # break
        print('Number of tweets found: {}'.format(len(response.data)),'Number of user found:{}'.format(len(response.includes['users'])))

        try:
            for user in response.includes['users']:
                user_dict[user.id] = {'username': user.username if user.username else None,
                                      'name':user.name if user.name else None,
                                      'profile_image_url':user.profile_image_url if user.profile_image_url else None,
                                      'followers_count': user.public_metrics['followers_count'] if user.public_metrics['followers_count'] else None,
                                      'following_count': user.public_metrics['following_count'] if user.public_metrics['following_count'] else None,
                                      'listed_count': user.public_metrics['listed_count'] if user.public_metrics['listed_count'] else None,
                                      'tweet_count': user.public_metrics['tweet_count'] if user.public_metrics['tweet_count'] else None,
                                      'description': user.description if user.description else None,
                                      'location': user.location if user.location else None,
                                      'url':user.url if user.url else None,
                                      'verified':user.verified if user.verified else None
                                      }
                # print(user_dict)
                user_df = pd.DataFrame(user_dict[user.id], index=[0])
                user_df['author_id'] =user.id
                user_df['author_id'] = user_df['author_id'].astype(str)

                author_profile = pd.concat([author_profile,user_df])
                # if save_type == 'excel':
                #     if os.path.isfile(file_path+'/{}.xlsx'.format('phd_author_profile')) == False:
                #         with pd.ExcelWriter(file_path+'/{}.xlsx'.format('phd_author_profile')) as writer:
                #             user_df.to_excel(writer)
                #
                #     else:
                #         # existing_user = pd.read_excel(file_path+'/{}.xlsx'.format('author_profile'))
                #         existing_user = pd.read_excel(file_path + '/{}.xlsx'.format('phd_author_profile'), dtype=str)
                #         if user_df.author_id.tolist()[0] not in existing_user.author_id.tolist():
                #             path = file_path+'/{}.xlsx'.format('phd_author_profile')
                #             sheet_name = 'Sheet1'
                #             wb = load_workbook(path)
                #             sheet = wb[sheet_name]
                #             # Find last row of written data
                #             row = sheet.max_row + 1
                #             for no, col in enumerate(user_df.columns.to_list()):
                #                 sheet.cell(row=row, column=no + 1).value = user_df[col].tolist()[0]
                #             wb.save(path)
                #             print('Author {}'.format(user_df.author_id.tolist()[0]) +' saved')
                #         else:
                #             print('Author {}'.format(user_df.author_id.tolist()[0]) + ' Already existed')
                if save_type=='csv':
                    user_df['author_id'] = user_df['author_id'].map(lambda x: 'a'+x+'b')
                    if os.path.isfile(file_path+'/{}.csv'.format(user_filename)) == False:
                        user_df.to_csv(file_path+'/{}.csv'.format(user_filename))

                    else:
                        existing_user = pd.read_csv(file_path+'/{}.csv'.format(user_filename), dtype=str,lineterminator='\n')
                        if user_df.author_id.tolist()[0] not in existing_user.author_id.tolist():
                            user_df.to_csv(file_path+'/{}.csv'.format(user_filename),mode='a',header=False)

        except:
            pass

        # try:
        #     for place in response.includes['places']:
        #         place_dict[place.id] = place.full_name
        #         break
        # except:
        #     pass

            # For each tweet, find the author's information
        try:
            for tweet in response.data:
                    result = []
                    result.append({
                        'id': tweet.id if tweet.id else None,
                        'author_id': tweet.author_id if tweet.author_id else None,
                        'text': tweet.text if tweet.text else None,
                        'created_at': tweet.created_at if tweet.created_at else None,
                        'retweet_count': tweet.public_metrics['retweet_count'] if tweet.public_metrics['retweet_count'] else None,
                        'reply_count': tweet.public_metrics['reply_count'] if tweet.public_metrics['reply_count'] else None,
                        'like_count': tweet.public_metrics['like_count'] if tweet.public_metrics['like_count'] else None,
                        'quote_count': tweet.public_metrics['quote_count'] if tweet.public_metrics['quote_count'] else None,
                        'geo': tweet.geo if tweet.geo else None,
                        'lang':tweet.lang if tweet.lang else None,
                        'conversation_id':tweet.conversation_id if tweet.conversation_id else None,
                        'referenced_tweets':{p['id']: p['type'] for p in tweet.referenced_tweets}
                                    if tweet.referenced_tweets else None},

                    )

                    df = pd.DataFrame(result)
                    df['keyword'] = keyword1+' '+ str(keyword2)
                    df['referenced_tweets'] = df['referenced_tweets'].astype(str)
                    df['id'] = df['id'].astype(str)
                    df['author_id'] = df['author_id'].astype(str)
                    df['conversation_id'] = df['conversation_id'].astype(str)
                    df['created_at'] = df['created_at'].apply(lambda a: pd.to_datetime(a).date())
                    bri = pd.concat([bri, df])
                    # if save_type =='excel':
                    #     if os.path.isfile(file_path+'/{}.xlsx'.format('brism_dataset1'))==False:
                    #         with pd.ExcelWriter(file_path+'/{}.xlsx'.format('brism_dataset1')) as writer:
                    #             df.to_excel(writer)
                    #             # print('saved')
                    #     else:
                    #         existing = pd.read_excel(file_path+'/{}.xlsx'.format('brism_dataset1'),dtype=str)
                    #
                    #         if df.id.tolist()[0] not in existing.id.tolist():
                    #             path =file_path+'/{}.xlsx'.format('brism_dataset1')
                    #             sheet_name = 'Sheet1'
                    #             wb = load_workbook(path)
                    #             sheet = wb[sheet_name]
                    #             # Find last row of written data
                    #             row = sheet.max_row + 1
                    #             for no, col in enumerate(df.columns.to_list()):
                    #                 sheet.cell(row=row, column=no+1).value = df[col].tolist()[0]
                    #             wb.save(path)
                    #             time.sleep(1)
                    #             print('Saved tweet {}'.format(df.id.tolist()[0]))
                    #         else:
                    #             print('Tweet {}'.format(df.id.tolist()[0])+' Already existed')

                    if save_type == 'csv':
                            df['id'] = df['id'].map(lambda x: 'a'+x+'b')
                            df['author_id'] = df['author_id'].map(lambda x: 'a'+x+'b')
                            df['conversation_id'] = df['conversation_id'].map(lambda x: 'a'+x+'b')
                            if os.path.isfile(file_path+'/{}.csv'.format(tweet_filename)) == False:
                                df.to_csv(file_path+'/{}.csv'.format(tweet_filename))

                            else:
                                existing = pd.read_csv(file_path+'/{}.csv'.format(tweet_filename), dtype=str,lineterminator='\n')
                                if df.id.tolist()[0] not in existing.id.tolist():
                                    df.to_csv(file_path+'/{}.csv'.format(tweet_filename), mode='a', header=False)
        except:
            pass

        print('Getting next page {}'.format(i+1))
        print('Latested timestamp: {}'.format(bri.created_at.iloc[-1]))
        time.sleep(1)
        i +=1
    # return belt_and_road,author_profile
def main():
    scrapy(keyword1 = 'maritime silk road',keyword2 = None,and_query =False, until_id = '1096251958938091522', save_type = 'csv',
           start = '2018-01-01',end='2021-12-31',file_path = '/Users/jie/phd_project/TwitterData/brism',
           user_filename = 'brism_users', tweet_filename = 'brism_tweets')
#keyword1 = 'maritime silk road”, and “silk road economic belt
#一带一路
if __name__ == "__main__":
    main()

    #961888225106120704
#keywords: belt and road scholarship; belt and road guangxi; belt and road exchange;belt and road student